"""Persistent target-company catalog, import, matching, and direct career search."""

from __future__ import annotations

from datetime import date, datetime, timedelta
from contextlib import contextmanager
from hashlib import sha256
from html import unescape
from html.parser import HTMLParser
from io import BytesIO
import json
from pathlib import Path
import re
import sqlite3
import unicodedata
from urllib.parse import parse_qsl, urlencode, urljoin, urlparse, urlunparse
from urllib.request import Request, urlopen


ROOT = Path(__file__).parent
DEFAULT_DB = ROOT / "output" / "target_companies.sqlite3"
SEED_FILE = ROOT / "data" / "target_companies_seed.json"

SENIOR_TITLE_RE = re.compile(
    r"\b(senior|sr\.?|director|head|vice president|vp|principal|chief)\b",
    re.I,
)

ROLE_FAMILIES = {
    "analytics": {
        "analyst", "analytics", "data", "insights", "intelligence", "bi",
        "reporting", "measurement",
    },
    "operations": {
        "operations", "operation", "operational", "ops", "bizops", "revops",
        "salesops", "execution", "process",
    },
    "strategy": {
        "strategy", "strategic", "planning", "corporate", "business",
    },
    "finance": {
        "finance", "financial", "fp&a", "fpa", "investment", "portfolio",
        "credit", "revenue", "pricing",
    },
    "product": {"product", "technical", "systems", "solutions", "platform"},
    "marketing": {
        "marketing", "growth", "brand", "content", "customer", "commercial",
        "marketplace",
    },
    "program": {
        "program", "project", "implementation", "transformation", "change",
    },
    "supply_chain": {
        "supply", "chain", "logistics", "inventory", "manufacturing",
        "procurement",
    },
    "risk": {
        "risk", "compliance", "legal", "controls", "governance", "fraud",
    },
    "healthcare": {
        "healthcare", "clinical", "biostatistical", "bioinformatics", "epic",
    },
}

ROLE_STOPWORDS = {
    "and", "or", "the", "of", "for", "to", "in", "with", "a", "an",
    "entry", "level", "junior", "assistant",
}


def _now() -> str:
    return datetime.now().replace(microsecond=0).isoformat()


def normalize_text(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.casefold().replace("&", " and ")
    return re.sub(r"[^a-z0-9+#]+", " ", text).strip()


def normalize_company(value: str) -> str:
    def parenthetical(match: re.Match) -> str:
        value = match.group(1).strip()
        if value.isupper() and len(value) <= 8:
            return f" {value} "
        return " "

    text = re.sub(r"\(([^)]*)\)", parenthetical, str(value or ""))
    normalized = normalize_text(text)
    words = [
        word for word in normalized.split()
        if word not in {
            "inc", "incorporated", "corp", "corporation", "company", "co",
            "llc", "ltd", "limited", "plc",
        }
    ]
    return " ".join(words)


def normalize_role(value: str) -> str:
    replacements = {
        "bizops": "business operations",
        "revops": "revenue operations",
        "sales ops": "sales operations",
        "data ops": "data operations",
        "strategy ops": "strategy operations",
        "strategic": "strategy",
        "operational": "operations",
        "analytics": "analyst",
    }
    text = f" {normalize_text(value)} "
    for old, new in replacements.items():
        text = text.replace(f" {old} ", f" {new} ")
    return re.sub(r"\s+", " ", text).strip()


def split_target_roles(value: str) -> list[str]:
    roles = []
    for comma_part in re.split(r"[,;\n]+", str(value or "")):
        comma_part = comma_part.strip()
        if not comma_part:
            continue
        slash_parts = re.split(r"\s+/\s+|(?<=\w)/(?=\w)", comma_part)
        for part in slash_parts:
            part = re.sub(r"\s+", " ", part).strip(" -")
            if part and normalize_role(part) not in {
                normalize_role(role) for role in roles
            }:
                roles.append(part)
    return roles


def _column_index(headers: list[str], terms: tuple[str, ...]) -> int | None:
    normalized = [normalize_text(value) for value in headers]
    for index, header in enumerate(normalized):
        if any(term in header for term in terms):
            return index
    return None


def records_from_workbook(content: bytes) -> list[dict]:
    """Read all sheets from the uploaded target-company workbook."""
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    records = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(value or "").strip() for value in rows[0]]
        company_col = _column_index(headers, ("company name", "company"))
        role_col = _column_index(headers, ("role categories", "roles", "role"))
        location_col = _column_index(headers, ("location",))
        career_col = _column_index(
            headers, ("career website", "career site", "careers url", "career url")
        )
        category_col = _column_index(headers, ("type", "profile", "category"))
        sponsorship_col = _column_index(
            headers, ("visa", "work authorization", "sponsorship")
        )
        if company_col is None or role_col is None:
            continue

        for values in rows[1:]:
            company = str(values[company_col] or "").strip()
            role_text = str(values[role_col] or "").strip()
            if not company:
                continue
            records.append({
                "company_name": company,
                "aliases": [company],
                "location": (
                    str(values[location_col] or "").strip()
                    if location_col is not None else ""
                ),
                "roles": split_target_roles(role_text),
                "career_url": (
                    str(values[career_col] or "").strip()
                    if career_col is not None else ""
                ),
                "category": (
                    str(values[category_col] or "").strip()
                    if category_col is not None else ""
                ),
                "source_tabs": [sheet.title],
                "notes": (
                    str(values[sponsorship_col] or "").strip()
                    if sponsorship_col is not None else ""
                ),
            })
    return records


@contextmanager
def _connect(db_path: Path | str = DEFAULT_DB):
    path = Path(db_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(path, timeout=30)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    try:
        yield connection
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


def initialize_database(
    db_path: Path | str = DEFAULT_DB,
    seed_path: Path | str = SEED_FILE,
) -> None:
    with _connect(db_path) as connection:
        connection.executescript(
            """
            CREATE TABLE IF NOT EXISTS target_companies (
                id INTEGER PRIMARY KEY,
                company_name TEXT NOT NULL,
                normalized_name TEXT NOT NULL UNIQUE,
                aliases_json TEXT NOT NULL DEFAULT '[]',
                location TEXT NOT NULL DEFAULT '',
                career_url TEXT NOT NULL DEFAULT '',
                category TEXT NOT NULL DEFAULT '',
                source_tabs_json TEXT NOT NULL DEFAULT '[]',
                notes TEXT NOT NULL DEFAULT '',
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                last_searched_at TEXT NOT NULL DEFAULT '',
                last_search_status TEXT NOT NULL DEFAULT '',
                last_job_count INTEGER NOT NULL DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS target_roles (
                id INTEGER PRIMARY KEY,
                company_id INTEGER NOT NULL REFERENCES target_companies(id)
                    ON DELETE CASCADE,
                role_name TEXT NOT NULL,
                normalized_role TEXT NOT NULL,
                UNIQUE(company_id, normalized_role)
            );
            CREATE TABLE IF NOT EXISTS search_runs (
                id INTEGER PRIMARY KEY,
                started_at TEXT NOT NULL,
                finished_at TEXT NOT NULL DEFAULT '',
                geographic_scope TEXT NOT NULL DEFAULT '',
                date_posted TEXT NOT NULL DEFAULT 'any',
                sources_json TEXT NOT NULL DEFAULT '[]',
                companies_total INTEGER NOT NULL DEFAULT 0,
                companies_succeeded INTEGER NOT NULL DEFAULT 0,
                companies_failed INTEGER NOT NULL DEFAULT 0,
                jobs_found INTEGER NOT NULL DEFAULT 0,
                jobs_new INTEGER NOT NULL DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS seen_jobs (
                id INTEGER PRIMARY KEY,
                job_key TEXT NOT NULL UNIQUE,
                fingerprint TEXT NOT NULL,
                provider_id TEXT NOT NULL DEFAULT '',
                canonical_url TEXT NOT NULL DEFAULT '',
                company_name TEXT NOT NULL,
                normalized_company TEXT NOT NULL,
                title TEXT NOT NULL,
                normalized_title TEXT NOT NULL,
                location TEXT NOT NULL DEFAULT '',
                source TEXT NOT NULL DEFAULT '',
                sources_json TEXT NOT NULL DEFAULT '[]',
                matched_company_id INTEGER REFERENCES target_companies(id)
                    ON DELETE SET NULL,
                matched_role TEXT NOT NULL DEFAULT '',
                date_posted TEXT NOT NULL DEFAULT '',
                description TEXT NOT NULL DEFAULT '',
                first_seen TEXT NOT NULL,
                last_seen TEXT NOT NULL,
                first_run_id INTEGER REFERENCES search_runs(id),
                last_run_id INTEGER REFERENCES search_runs(id),
                times_seen INTEGER NOT NULL DEFAULT 1
            );
            CREATE INDEX IF NOT EXISTS idx_seen_jobs_fingerprint
                ON seen_jobs(fingerprint);
            CREATE INDEX IF NOT EXISTS idx_seen_jobs_last_run
                ON seen_jobs(last_run_id);
            """
        )
        columns = {
            row[1] for row in connection.execute("PRAGMA table_info(seen_jobs)")
        }
        if "first_run_id" not in columns:
            connection.execute(
                "ALTER TABLE seen_jobs ADD COLUMN first_run_id INTEGER"
            )
        count = connection.execute(
            "SELECT COUNT(*) FROM target_companies"
        ).fetchone()[0]

    seed = Path(seed_path)
    if count == 0 and seed.exists():
        import_records(json.loads(seed.read_text(encoding="utf-8")), db_path)


def _merge_text(existing: str, incoming: str) -> str:
    values = []
    for value in (existing, incoming):
        for item in re.split(r"\s*[|]\s*", value or ""):
            item = item.strip()
            if item and normalize_text(item) not in {
                normalize_text(current) for current in values
            }:
                values.append(item)
    return " | ".join(values)


def _merge_json(existing: str, incoming: list[str]) -> str:
    values = json.loads(existing or "[]")
    for item in incoming:
        if item and normalize_text(item) not in {
            normalize_text(current) for current in values
        }:
            values.append(item)
    return json.dumps(values, ensure_ascii=False)


def import_records(
    records: list[dict],
    db_path: Path | str = DEFAULT_DB,
) -> dict:
    """Idempotently merge imported companies, roles, and missing career URLs."""
    initialize_database(db_path, seed_path=Path("__no_seed__"))
    result = {
        "companies_added": 0,
        "companies_updated": 0,
        "roles_added": 0,
        "duplicates_merged": 0,
        "career_urls_added": 0,
        "career_url_conflicts": [],
    }
    now = _now()
    with _connect(db_path) as connection:
        for record in records:
            company_name = str(record.get("company_name") or "").strip()
            normalized = normalize_company(company_name)
            if not normalized:
                continue
            row = connection.execute(
                "SELECT * FROM target_companies WHERE normalized_name = ?",
                (normalized,),
            ).fetchone()
            incoming_url = str(record.get("career_url") or "").strip()
            if row is None:
                cursor = connection.execute(
                    """
                    INSERT INTO target_companies (
                        company_name, normalized_name, aliases_json, location,
                        career_url, category, source_tabs_json, notes, active,
                        created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
                    """,
                    (
                        company_name,
                        normalized,
                        json.dumps(record.get("aliases") or [company_name],
                                   ensure_ascii=False),
                        str(record.get("location") or "").strip(),
                        incoming_url,
                        str(record.get("category") or "").strip(),
                        json.dumps(record.get("source_tabs") or [],
                                   ensure_ascii=False),
                        str(record.get("notes") or "").strip(),
                        now,
                        now,
                    ),
                )
                company_id = cursor.lastrowid
                result["companies_added"] += 1
                if incoming_url:
                    result["career_urls_added"] += 1
            else:
                company_id = row["id"]
                aliases = list(record.get("aliases") or []) + [company_name]
                career_url = row["career_url"]
                if not career_url and incoming_url:
                    career_url = incoming_url
                    result["career_urls_added"] += 1
                elif (
                    career_url and incoming_url
                    and canonicalize_url(career_url) != canonicalize_url(incoming_url)
                ):
                    result["career_url_conflicts"].append({
                        "company": row["company_name"],
                        "existing": career_url,
                        "imported": incoming_url,
                    })
                connection.execute(
                    """
                    UPDATE target_companies
                    SET aliases_json = ?, location = ?, career_url = ?,
                        category = ?, source_tabs_json = ?, notes = ?,
                        updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        _merge_json(row["aliases_json"], aliases),
                        _merge_text(row["location"],
                                    str(record.get("location") or "").strip()),
                        career_url,
                        _merge_text(row["category"],
                                    str(record.get("category") or "").strip()),
                        _merge_json(
                            row["source_tabs_json"],
                            list(record.get("source_tabs") or []),
                        ),
                        _merge_text(row["notes"],
                                    str(record.get("notes") or "").strip()),
                        now,
                        company_id,
                    ),
                )
                result["companies_updated"] += 1
                result["duplicates_merged"] += 1

            for role in record.get("roles") or []:
                normalized_role = normalize_role(role)
                if not normalized_role:
                    continue
                before = connection.total_changes
                connection.execute(
                    """
                    INSERT OR IGNORE INTO target_roles
                        (company_id, role_name, normalized_role)
                    VALUES (?, ?, ?)
                    """,
                    (company_id, role.strip(), normalized_role),
                )
                if connection.total_changes > before:
                    result["roles_added"] += 1
    return result


def list_companies(
    db_path: Path | str = DEFAULT_DB,
    active_only: bool = False,
) -> list[dict]:
    initialize_database(db_path)
    where = "WHERE c.active = 1" if active_only else ""
    with _connect(db_path) as connection:
        rows = connection.execute(
            f"""
            SELECT c.*, GROUP_CONCAT(r.role_name, ' || ') AS roles_text
            FROM target_companies c
            LEFT JOIN target_roles r ON r.company_id = c.id
            {where}
            GROUP BY c.id
            ORDER BY c.company_name COLLATE NOCASE
            """
        ).fetchall()
    result = []
    for row in rows:
        item = dict(row)
        item["roles"] = [
            role for role in (item.pop("roles_text") or "").split(" || ") if role
        ]
        item["aliases"] = json.loads(item.pop("aliases_json") or "[]")
        item["source_tabs"] = json.loads(item.pop("source_tabs_json") or "[]")
        item["active"] = bool(item["active"])
        result.append(item)
    return result


def save_company(
    company_name: str,
    roles: list[str],
    location: str = "",
    career_url: str = "",
    category: str = "",
    notes: str = "",
    active: bool = True,
    company_id: int | None = None,
    db_path: Path | str = DEFAULT_DB,
) -> int:
    initialize_database(db_path)
    normalized = normalize_company(company_name)
    if not normalized:
        raise ValueError("Company name is required.")
    now = _now()
    with _connect(db_path) as connection:
        if company_id is None:
            existing = connection.execute(
                "SELECT id FROM target_companies WHERE normalized_name = ?",
                (normalized,),
            ).fetchone()
            if existing:
                company_id = existing["id"]
            else:
                cursor = connection.execute(
                    """
                    INSERT INTO target_companies (
                        company_name, normalized_name, aliases_json, location,
                        career_url, category, notes, active, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        company_name.strip(), normalized,
                        json.dumps([company_name.strip()], ensure_ascii=False),
                        location.strip(), career_url.strip(), category.strip(),
                        notes.strip(), int(active), now, now,
                    ),
                )
                company_id = cursor.lastrowid
        connection.execute(
            """
            UPDATE target_companies
            SET company_name = ?, normalized_name = ?, location = ?,
                career_url = ?, category = ?, notes = ?, active = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (
                company_name.strip(), normalized, location.strip(),
                career_url.strip(), category.strip(), notes.strip(), int(active),
                now, company_id,
            ),
        )
        connection.execute(
            "DELETE FROM target_roles WHERE company_id = ?", (company_id,)
        )
        for role in roles:
            role = role.strip()
            if role:
                connection.execute(
                    """
                    INSERT OR IGNORE INTO target_roles
                        (company_id, role_name, normalized_role)
                    VALUES (?, ?, ?)
                    """,
                    (company_id, role, normalize_role(role)),
                )
    return int(company_id)


def delete_companies(
    company_ids: list[int],
    db_path: Path | str = DEFAULT_DB,
) -> None:
    if not company_ids:
        return
    placeholders = ",".join("?" for _ in company_ids)
    with _connect(db_path) as connection:
        connection.execute(
            f"""
            UPDATE seen_jobs SET matched_company_id = NULL
            WHERE matched_company_id IN ({placeholders})
            """,
            tuple(company_ids),
        )
        connection.execute(
            f"DELETE FROM target_companies WHERE id IN ({placeholders})",
            tuple(company_ids),
        )


def role_match_score(title: str, target_roles: list[str]) -> tuple[int, str]:
    if not title or SENIOR_TITLE_RE.search(title):
        return 0, ""
    normalized_title = normalize_role(title)
    title_tokens = set(normalized_title.split()) - ROLE_STOPWORDS
    best = (0, "")
    for target_role in target_roles:
        normalized_target = normalize_role(target_role)
        target_tokens = set(normalized_target.split()) - ROLE_STOPWORDS
        if not target_tokens:
            continue
        if normalized_target in normalized_title or normalized_title in normalized_target:
            score = 100
        else:
            overlap = title_tokens & target_tokens
            union = title_tokens | target_tokens
            lexical = int(70 * len(overlap) / max(1, len(union)))
            family_bonus = 0
            for terms in ROLE_FAMILIES.values():
                if title_tokens & terms and target_tokens & terms:
                    family_bonus += 18
            level_bonus = 12 if (
                {"analyst", "consultant", "specialist", "associate", "manager"}
                & title_tokens
                & target_tokens
            ) else 0
            score = min(95, lexical + family_bonus + level_bonus)
        if score > best[0]:
            best = (score, target_role)
    return best


def company_matches(job_company: str, target_company: dict) -> bool:
    job = normalize_company(job_company)
    names = [target_company["company_name"]] + target_company.get("aliases", [])
    normalized_names = {normalize_company(name) for name in names}
    if not job:
        return False
    job_tokens = set(job.split())
    for name in normalized_names:
        if not name:
            continue
        name_tokens = set(name.split())
        if (
            job == name
            or (len(name) >= 5 and (name in job or job in name))
            or (name_tokens and name_tokens <= job_tokens)
            or (
                len(name_tokens) >= 2
                and len(name_tokens & job_tokens) / len(name_tokens) >= 0.75
            )
        ):
            return True
    return False


def geography_matches(job_location: str, scope: str) -> bool:
    if not scope.strip():
        return True
    job = normalize_text(job_location)
    requested = normalize_text(scope)
    if not job:
        return True
    if "remote" in requested and "remote" in job:
        return True
    if requested in job or job in requested:
        return True
    tokens = set(requested.split()) - {"usa", "us", "united", "states", "area"}
    if tokens & set(job.split()):
        return True
    if "california" in tokens and (
        re.search(r"\bca\b", job)
        or any(city in job for city in (
            "san francisco", "san jose", "oakland", "palo alto", "sunnyvale",
            "santa clara", "mountain view", "redwood city", "foster city",
            "cupertino", "menlo park", "pleasanton", "san mateo", "fremont",
        ))
    ):
        return True
    return False


def canonicalize_url(value: str) -> str:
    if not value:
        return ""
    parsed = urlparse(value.strip())
    ignored = {
        "trk", "trackingid", "ref", "refid", "source", "src", "gh_src",
        "lever-source", "utm_source", "utm_medium", "utm_campaign",
        "utm_content", "utm_term",
    }
    query = [
        (key, val) for key, val in parse_qsl(parsed.query, keep_blank_values=True)
        if key.casefold() not in ignored
    ]
    return urlunparse((
        parsed.scheme.casefold() or "https",
        parsed.netloc.casefold(),
        re.sub(r"/+$", "", parsed.path),
        "",
        urlencode(sorted(query)),
        "",
    ))


def job_keys(job: dict) -> tuple[str, str]:
    company = normalize_company(job.get("company", ""))
    title = normalize_role(job.get("title", ""))
    location = normalize_text(job.get("location", ""))
    source = normalize_text(job.get("board") or job.get("source") or "")
    provider_id = normalize_text(str(job.get("id") or ""))
    url = canonicalize_url(job.get("url", ""))
    if provider_id:
        stable = f"{source}:{provider_id}"
    elif url:
        stable = url
    else:
        stable = f"{company}|{title}|{location}"
    fingerprint = f"{company}|{title}|{location}"
    return (
        sha256(stable.encode("utf-8")).hexdigest(),
        sha256(fingerprint.encode("utf-8")).hexdigest(),
    )


class _JobLinkParser(HTMLParser):
    def __init__(self, base_url: str):
        super().__init__()
        self.base_url = base_url
        self.links: list[tuple[str, str]] = []
        self._href = ""
        self._text: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]):
        if tag.casefold() == "a":
            self._href = dict(attrs).get("href") or ""
            self._text = []

    def handle_data(self, data: str):
        if self._href:
            self._text.append(data)

    def handle_endtag(self, tag: str):
        if tag.casefold() == "a" and self._href:
            text = re.sub(r"\s+", " ", " ".join(self._text)).strip()
            href = urljoin(self.base_url, self._href)
            if text and href:
                self.links.append((text, href))
            self._href = ""
            self._text = []


def _request(url: str, timeout: int = 30) -> bytes:
    request = Request(url, headers={
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 Chrome/131.0.0.0 Safari/537.36"
        ),
        "Accept": "application/json,text/html;q=0.9,*/*;q=0.8",
    })
    with urlopen(request, timeout=timeout) as response:
        return response.read()


def _greenhouse_jobs(company: str, career_url: str) -> list[dict] | None:
    parsed = urlparse(career_url)
    if "greenhouse.io" not in parsed.netloc.casefold():
        return None
    parts = [part for part in parsed.path.split("/") if part]
    if not parts:
        return None
    token = parts[0] if parts[0] != "embed" else ""
    if not token:
        token = dict(parse_qsl(parsed.query)).get("for", "")
    if not token:
        return None
    data = json.loads(_request(
        f"https://boards-api.greenhouse.io/v1/boards/{token}/jobs?content=true"
    ))
    return [{
        "id": str(job.get("id", "")),
        "title": job.get("title", ""),
        "company": company,
        "location": (job.get("location") or {}).get("name", ""),
        "url": job.get("absolute_url", ""),
        "description": unescape(job.get("content", "")),
        "date_posted": job.get("updated_at", ""),
        "board": "Career site",
        "source": "Greenhouse",
    } for job in data.get("jobs", [])]


def _lever_jobs(company: str, career_url: str) -> list[dict] | None:
    parsed = urlparse(career_url)
    host = parsed.netloc.casefold()
    if "lever.co" not in host:
        return None
    parts = [part for part in parsed.path.split("/") if part]
    if not parts:
        return None
    site = parts[0]
    api_host = "api.eu.lever.co" if ".eu.lever.co" in host else "api.lever.co"
    data = json.loads(_request(
        f"https://{api_host}/v0/postings/{site}?mode=json"
    ))
    return [{
        "id": str(job.get("id", "")),
        "title": job.get("text", ""),
        "company": company,
        "location": (job.get("categories") or {}).get("location", ""),
        "url": job.get("hostedUrl", ""),
        "description": job.get("descriptionPlain", ""),
        "date_posted": "",
        "board": "Career site",
        "source": "Lever",
    } for job in data]


def _smartrecruiters_jobs(company: str, career_url: str) -> list[dict] | None:
    parsed = urlparse(career_url)
    if "smartrecruiters.com" not in parsed.netloc.casefold():
        return None
    parts = [part for part in parsed.path.split("/") if part]
    if not parts:
        return None
    identifier = parts[0]
    data = json.loads(_request(
        "https://api.smartrecruiters.com/v1/companies/"
        f"{identifier}/postings?limit=100"
    ))
    jobs = []
    for job in data.get("content", []):
        location = job.get("location") or {}
        location_text = ", ".join(
            str(location.get(key) or "") for key in ("city", "region", "country")
            if location.get(key)
        )
        jobs.append({
            "id": str(job.get("id", "")),
            "title": job.get("name", ""),
            "company": company,
            "location": location_text,
            "url": job.get("ref", "") or (
                f"https://jobs.smartrecruiters.com/{identifier}/{job.get('id', '')}"
            ),
            "description": "",
            "date_posted": job.get("releasedDate", ""),
            "board": "Career site",
            "source": "SmartRecruiters",
        })
    return jobs


def fetch_career_jobs(
    company: str,
    career_url: str,
    timeout: int = 30,
) -> tuple[list[dict], str, str]:
    """Fetch published jobs from a supported ATS or a generic career page."""
    if not career_url.strip():
        return [], "missing_url", "No career URL"
    try:
        for provider in (_greenhouse_jobs, _lever_jobs, _smartrecruiters_jobs):
            jobs = provider(company, career_url)
            if jobs is not None:
                return jobs, "ok", provider.__name__.strip("_").replace("_jobs", "")

        raw = _request(career_url, timeout=timeout).decode(
            "utf-8", errors="replace"
        )
        parser = _JobLinkParser(career_url)
        parser.feed(raw)
        jobs = []
        seen_urls = set()
        for text, url in parser.links:
            normalized_url = canonicalize_url(url)
            combined = normalize_text(f"{text} {url}")
            if (
                normalized_url in seen_urls
                or not re.search(
                    r"\b(job|jobs|career|careers|position|positions|opening|"
                    r"greenhouse|lever|workday|ashby|smartrecruiters)\b",
                    combined,
                )
            ):
                continue
            seen_urls.add(normalized_url)
            jobs.append({
                "id": "",
                "title": text,
                "company": company,
                "location": "",
                "url": url,
                "description": "",
                "date_posted": "",
                "board": "Career site",
                "source": "Career page",
            })
        if jobs:
            return jobs, "ok", "Generic career page"
        return [], "unsupported", "Page loaded but no job links were readable"
    except Exception as error:
        return [], "error", str(error)[:240]


def start_search_run(
    geographic_scope: str,
    date_posted: str,
    sources: list[str],
    companies_total: int,
    db_path: Path | str = DEFAULT_DB,
) -> int:
    initialize_database(db_path)
    with _connect(db_path) as connection:
        cursor = connection.execute(
            """
            INSERT INTO search_runs (
                started_at, geographic_scope, date_posted, sources_json,
                companies_total
            ) VALUES (?, ?, ?, ?, ?)
            """,
            (
                _now(), geographic_scope, date_posted,
                json.dumps(sources, ensure_ascii=False), companies_total,
            ),
        )
        return int(cursor.lastrowid)


def record_job(
    run_id: int,
    job: dict,
    target_company: dict,
    matched_role: str,
    db_path: Path | str = DEFAULT_DB,
) -> tuple[int, bool]:
    job_key, fingerprint = job_keys(job)
    today = str(date.today())
    source = str(job.get("source") or job.get("board") or "")
    canonical_url = canonicalize_url(job.get("url", ""))
    with _connect(db_path) as connection:
        existing = connection.execute(
            "SELECT * FROM seen_jobs WHERE job_key = ?", (job_key,)
        ).fetchone()
        if existing is None:
            cutoff = str(date.today() - timedelta(days=60))
            existing = connection.execute(
                """
                SELECT * FROM seen_jobs
                WHERE fingerprint = ? AND last_seen >= ?
                ORDER BY last_seen DESC LIMIT 1
                """,
                (fingerprint, cutoff),
            ).fetchone()
        if existing is not None:
            sources = json.loads(existing["sources_json"] or "[]")
            if source and source not in sources:
                sources.append(source)
            connection.execute(
                """
                UPDATE seen_jobs
                SET last_seen = ?, last_run_id = ?, times_seen = times_seen + 1,
                    sources_json = ?, canonical_url = CASE
                        WHEN canonical_url = '' THEN ? ELSE canonical_url END,
                    description = CASE
                        WHEN description = '' THEN ? ELSE description END
                WHERE id = ?
                """,
                (
                    today, run_id, json.dumps(sources, ensure_ascii=False),
                    canonical_url, str(job.get("description") or ""),
                    existing["id"],
                ),
            )
            return int(existing["id"]), False

        cursor = connection.execute(
            """
            INSERT INTO seen_jobs (
                job_key, fingerprint, provider_id, canonical_url,
                company_name, normalized_company, title, normalized_title,
                location, source, sources_json, matched_company_id,
                matched_role, date_posted, description, first_seen, last_seen,
                first_run_id, last_run_id
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                job_key, fingerprint, str(job.get("id") or ""), canonical_url,
                str(job.get("company") or target_company["company_name"]),
                normalize_company(job.get("company") or target_company["company_name"]),
                str(job.get("title") or ""),
                normalize_role(job.get("title") or ""),
                str(job.get("location") or ""), source,
                json.dumps([source] if source else [], ensure_ascii=False),
                target_company["id"], matched_role,
                str(job.get("date_posted") or ""),
                str(job.get("description") or ""),
                today, today, run_id, run_id,
            ),
        )
        return int(cursor.lastrowid), True


def update_company_search_status(
    company_id: int,
    status: str,
    job_count: int,
    db_path: Path | str = DEFAULT_DB,
) -> None:
    with _connect(db_path) as connection:
        connection.execute(
            """
            UPDATE target_companies
            SET last_searched_at = ?, last_search_status = ?,
                last_job_count = ?
            WHERE id = ?
            """,
            (_now(), status, job_count, company_id),
        )


def finish_search_run(
    run_id: int,
    companies_succeeded: int,
    companies_failed: int,
    jobs_found: int,
    jobs_new: int,
    db_path: Path | str = DEFAULT_DB,
) -> None:
    with _connect(db_path) as connection:
        connection.execute(
            """
            UPDATE search_runs
            SET finished_at = ?, companies_succeeded = ?,
                companies_failed = ?, jobs_found = ?, jobs_new = ?
            WHERE id = ?
            """,
            (
                _now(), companies_succeeded, companies_failed,
                jobs_found, jobs_new, run_id,
            ),
        )


def run_jobs(
    run_id: int,
    new_only: bool = False,
    db_path: Path | str = DEFAULT_DB,
) -> list[dict]:
    initialize_database(db_path)
    where = "AND j.first_run_id = ?" if new_only else ""
    params = (run_id, run_id) if new_only else (run_id,)
    with _connect(db_path) as connection:
        rows = connection.execute(
            f"""
            SELECT j.*, c.company_name AS target_company
            FROM seen_jobs j
            JOIN search_runs r ON r.id = j.last_run_id
            LEFT JOIN target_companies c ON c.id = j.matched_company_id
            WHERE j.last_run_id = ? {where}
            ORDER BY j.company_name COLLATE NOCASE, j.title COLLATE NOCASE
            """,
            params,
        ).fetchall()
    return [dict(row) for row in rows]


def database_summary(db_path: Path | str = DEFAULT_DB) -> dict:
    initialize_database(db_path)
    with _connect(db_path) as connection:
        return {
            "companies": connection.execute(
                "SELECT COUNT(*) FROM target_companies"
            ).fetchone()[0],
            "active_companies": connection.execute(
                "SELECT COUNT(*) FROM target_companies WHERE active = 1"
            ).fetchone()[0],
            "roles": connection.execute(
                "SELECT COUNT(*) FROM target_roles"
            ).fetchone()[0],
            "career_urls": connection.execute(
                "SELECT COUNT(*) FROM target_companies WHERE career_url <> ''"
            ).fetchone()[0],
            "seen_jobs": connection.execute(
                "SELECT COUNT(*) FROM seen_jobs"
            ).fetchone()[0],
        }
